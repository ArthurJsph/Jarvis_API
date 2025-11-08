"""
Manipulação de Arquivos Excel
Jarvis CLI - Módulo de Excel
"""

import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from typing import List, Dict, Any, Optional
import config
from rich.console import Console
from rich.table import Table

console = Console()


class ExcelManager:
    """Gerenciador de arquivos Excel"""
    
    def __init__(self):
        self.encoding = config.DEFAULT_ENCODING
        self.date_format = config.DATE_FORMAT
        self.csv_separator = config.CSV_SEPARATOR
    
    def create_excel(self, filename: str, data: Optional[Dict[str, List]] = None, 
                    sheet_name: str = "Sheet1") -> bool:
        """
        Cria um novo arquivo Excel
        
        Args:
            filename: Nome do arquivo
            data: Dados para preencher (formato: {"coluna": [valores]})
            sheet_name: Nome da planilha
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if data:
                # Adiciona cabeçalhos
                headers = list(data.keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Adiciona dados
                max_rows = max(len(values) for values in data.values()) if data else 0
                for row_idx in range(max_rows):
                    for col_idx, header in enumerate(headers, 1):
                        if row_idx < len(data[header]):
                            ws.cell(row=row_idx + 2, column=col_idx, value=data[header][row_idx])
            
            # Salva o arquivo
            filepath = Path(config.DOCUMENTS_DIR) / filename
            wb.save(filepath)
            
            console.print(f"✅ Arquivo Excel criado: {filepath}", style="green")
            return True
            
        except Exception as e:
            console.print(f"❌ Erro ao criar Excel: {str(e)}", style="red")
            return False
    
    def read_excel(self, filename: str, sheet_name: Optional[str] = None) -> Optional[pd.DataFrame]:
        """
        Lê dados de um arquivo Excel
        
        Args:
            filename: Nome do arquivo
            sheet_name: Nome da planilha (None para a primeira)
        """
        try:
            filepath = Path(filename)
            if not filepath.exists():
                filepath = Path(config.DOCUMENTS_DIR) / filename
                
            if not filepath.exists():
                console.print(f"❌ Arquivo não encontrado: {filename}", style="red")
                return None
            
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            
            console.print(f"📊 Dados carregados do Excel: {filepath}", style="green")
            
            # Mostra preview dos dados
            table = Table(title=f"Preview: {filepath.name}")
            
            # Adiciona colunas
            for col in df.columns:
                table.add_column(str(col), style="cyan")
            
            # Adiciona algumas linhas (máximo 5)
            for idx, row in df.head().iterrows():
                table.add_row(*[str(val) for val in row])
            
            console.print(table)
            return df
            
        except Exception as e:
            console.print(f"❌ Erro ao ler Excel: {str(e)}", style="red")
            return None
    
    def update_excel(self, filename: str, updates: Dict[str, Any], 
                    row: Optional[int] = None, column: Optional[str] = None) -> bool:
        """
        Atualiza dados em um arquivo Excel
        
        Args:
            filename: Nome do arquivo
            updates: Dados para atualizar
            row: Linha específica para atualizar
            column: Coluna específica para atualizar
        """
        try:
            filepath = Path(filename)
            if not filepath.exists():
                filepath = Path(config.DOCUMENTS_DIR) / filename
                
            if not filepath.exists():
                console.print(f"❌ Arquivo não encontrado: {filename}", style="red")
                return False
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            if row and column:
                # Atualiza célula específica
                ws[f"{column}{row}"] = updates.get("value", "")
            else:
                # Atualização em massa
                df = pd.read_excel(filepath)
                for col, value in updates.items():
                    if col in df.columns:
                        df[col] = value
                
                # Salva de volta
                df.to_excel(filepath, index=False)
                console.print(f"✅ Dados atualizados em massa: {filepath}", style="green")
                return True
            
            wb.save(filepath)
            console.print(f"✅ Excel atualizado: {filepath}", style="green")
            return True
            
        except Exception as e:
            console.print(f"❌ Erro ao atualizar Excel: {str(e)}", style="red")
            return False
    
    def clean_excel(self, filename: str, remove_empty_rows: bool = True, 
                   remove_empty_cols: bool = True) -> bool:
        """
        Limpa dados desnecessários do Excel
        
        Args:
            filename: Nome do arquivo
            remove_empty_rows: Remove linhas vazias
            remove_empty_cols: Remove colunas vazias
        """
        try:
            filepath = Path(filename)
            if not filepath.exists():
                filepath = Path(config.DOCUMENTS_DIR) / filename
                
            df = pd.read_excel(filepath)
            original_shape = df.shape
            
            if remove_empty_rows:
                df = df.dropna(how='all')
            
            if remove_empty_cols:
                df = df.dropna(axis=1, how='all')
            
            # Remove colunas com nomes vazios
            df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
            
            df.to_excel(filepath, index=False)
            
            new_shape = df.shape
            console.print(f"✅ Excel limpo: {original_shape} → {new_shape}", style="green")
            return True
            
        except Exception as e:
            console.print(f"❌ Erro ao limpar Excel: {str(e)}", style="red")
            return False
    
    def excel_to_csv(self, excel_file: str, csv_file: Optional[str] = None, 
                    sheet_name: Optional[str] = None) -> bool:
        """
        Converte Excel para CSV
        
        Args:
            excel_file: Arquivo Excel de origem
            csv_file: Arquivo CSV de destino
            sheet_name: Planilha específica
        """
        try:
            excel_path = Path(excel_file)
            if not excel_path.exists():
                excel_path = Path(config.DOCUMENTS_DIR) / excel_file
            
            if not csv_file:
                csv_file = excel_path.stem + ".csv"
            
            csv_path = Path(config.DOCUMENTS_DIR) / csv_file
            
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            df.to_csv(csv_path, index=False, sep=self.csv_separator, encoding=self.encoding)
            
            console.print(f"✅ Convertido Excel → CSV: {csv_path}", style="green")
            return True
            
        except Exception as e:
            console.print(f"❌ Erro na conversão Excel → CSV: {str(e)}", style="red")
            return False
    
    def csv_to_excel(self, csv_file: str, excel_file: Optional[str] = None, 
                    sheet_name: str = "Sheet1") -> bool:
        """
        Converte CSV para Excel
        
        Args:
            csv_file: Arquivo CSV de origem
            excel_file: Arquivo Excel de destino
            sheet_name: Nome da planilha
        """
        try:
            csv_path = Path(csv_file)
            if not csv_path.exists():
                csv_path = Path(config.DOCUMENTS_DIR) / csv_file
            
            if not excel_file:
                excel_file = csv_path.stem + ".xlsx"
            
            excel_path = Path(config.DOCUMENTS_DIR) / excel_file
            
            df = pd.read_csv(csv_path, sep=self.csv_separator, encoding=self.encoding)
            df.to_excel(excel_path, index=False, sheet_name=sheet_name)
            
            console.print(f"✅ Convertido CSV → Excel: {excel_path}", style="green")
            return True
            
        except Exception as e:
            console.print(f"❌ Erro na conversão CSV → Excel: {str(e)}", style="red")
            return False
    
    def get_excel_info(self, filename: str) -> Dict[str, Any]:
        """
        Obtém informações sobre um arquivo Excel
        
        Args:
            filename: Nome do arquivo
        """
        try:
            filepath = Path(filename)
            if not filepath.exists():
                filepath = Path(config.DOCUMENTS_DIR) / filename
            
            wb = load_workbook(filepath)
            df = pd.read_excel(filepath)
            
            info = {
                "filename": filepath.name,
                "size": f"{filepath.stat().st_size / 1024:.2f} KB",
                "sheets": wb.sheetnames,
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": list(df.columns),
                "data_types": dict(df.dtypes),
                "null_counts": dict(df.isnull().sum())
            }
            
            return info
            
        except Exception as e:
            console.print(f"❌ Erro ao obter info do Excel: {str(e)}", style="red")
            return {}
    
    def generate_sample_data(self, filename: str, rows: int = 100) -> bool:
        """
        Gera dados de exemplo para testes
        
        Args:
            filename: Nome do arquivo
            rows: Número de linhas
        """
        try:
            from faker import Faker
            fake = Faker('pt_BR')
            
            data = {
                "ID": list(range(1, rows + 1)),
                "Nome": [fake.name() for _ in range(rows)],
                "Email": [fake.email() for _ in range(rows)],
                "Telefone": [fake.phone_number() for _ in range(rows)],
                "Cidade": [fake.city() for _ in range(rows)],
                "Data_Nascimento": [fake.date_of_birth().strftime(self.date_format) for _ in range(rows)],
                "Salario": [fake.pydecimal(left_digits=4, right_digits=2, positive=True) for _ in range(rows)]
            }
            
            return self.create_excel(filename, data, "Dados_Exemplo")
            
        except ImportError:
            console.print("❌ Faker não instalado. Use: pip install faker", style="red")
            return False
        except Exception as e:
            console.print(f"❌ Erro ao gerar dados de exemplo: {str(e)}", style="red")
            return False